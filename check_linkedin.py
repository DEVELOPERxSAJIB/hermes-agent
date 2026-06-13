import openpyxl
wb = openpyxl.load_workbook('/home/ubuntu/nanosoft/nanosoft.xlsx', read_only=True)
if 'LinkedIn' in wb.sheetnames:
    ws = wb['LinkedIn']
    rows = list(ws.iter_rows(values_only=True))
    print(f"Total rows (incl header): {len(rows)}")
    print(f"Data rows: {len(rows)-1}")
    if rows:
        print(f"Header: {rows[0]}")
else:
    print("No LinkedIn tab found")
    print(f"Available sheets: {wb.sheetnames}")
