import openpyxl
wb = openpyxl.load_workbook('nanosoft.xlsx', read_only=True)
if 'LinkedIn' in wb.sheetnames:
    ws = wb['LinkedIn']
    rows = list(ws.iter_rows(values_only=True))
    print(f'LinkedIn tab rows (including header): {len(rows)}')
    print(f'Header: {rows[0]}')
    if len(rows) > 1:
        print(f'Data rows: {len(rows) - 1}')
        print(f'Last row: {rows[-1]}')
else:
    print('No LinkedIn tab found')
    print(f'Available sheets: {wb.sheetnames}')
wb.close()
